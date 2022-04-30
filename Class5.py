#Exceptions
#ways to make only valid entries

try:
    age = int(input("what is your age?: "))
    print ("Your age is " + str(age))
except ValueError:
    print("Please enter a valid value")

#Applying to rock paper scissors
 #creating RockPaperScissors Game


import random
user_action = input('Enter a choice of rock, paper, or scissors: ')
possible_action = ['rock', 'paper', 'scissors']
computer_action = random.choice(possible_action)
print('You chose ' + user_action + '. Computer chose ' + computer_action + '.')

def apple():
    while True:
        try:
        except ValueError:
            print("you must enter rock, paper, scissors")
            break #stops running try except function

if user_action.lower() ==computer_action:
    print('Both players selected ' + user_action + '. It is a tie.')
elif user_action.lower() == 'rock':
    if computer_action == 'scissors':
        print('Rock smashes scissors. You win.')
    else:
        print('Paper covers rock. You lose.')

elif user_action.lower() == 'scissors':
    if computer_action == 'paper':
        print('Scissors cuts the papers. You win')
    else:
        print('Rock smashes scissors. You lose')

elif user_action.lower() == 'paper':
    if computer_action == 'rock':
        print('Paper covers the rock. You win')
    else:
        print('Scissors cut paper. You lose')


#for loop and while loop
#for loop - runs exactly a fixed number times
#while loop - runs upto certain or upto when condition is fufilled


fruit = ['grape', 'cherry', 'lemon']
for items in fruit:
    print(items)

#shorter way of writing
for items in ['grape', 'cherry', 'lemon']
    print(items)

#similar code
for items in [1, 2, 3]
    print(items)

for numbers in range(1,20):
    print(numbers)

#lists
prices = [10, 20, 30, 40, 50] #we call it list
total = 0
for numbers in prices:
    total = total+numbers
print(total)

#create a python project that allows to pick random names
import random
possible_name = ['Shwadhin', 'Phil', 'Amanda']
computer_action = random.choice(possible_name)
print("The computer has chosen your name " + computer_action + ". Answer the question")

#simialr code
import random
computer_action = random.randint(1, 19000)
print("this os the powerball number: " + str(computer_action) ) ". Answer this question"

#with a file xslx
from openpyxl import workbook, load_workbook
import random
wb = load_workbook("hello.xlsx")
ws = wb.active
rangeline = ws["A2": "A19"] #we are trying to create a list
#our list is named as rangeline above (we can name it anything)
name = [] #creating an empty variable ready to take list
for items in rangeline:             #this gave us cell value
    for subitems in items:           #we want whats inside the cell
        name.append(subitems.value)
print(name)
computer_action = random.choice(name)
print("the computer randomly chose: " + computer_action)

#final code
from tkinter import*
from tkinter.ttk import*
from openpyxl import workbook, load_workbook
import random
root = Tk()
root.title("Randomization lottery")
root.geometry("300x300")
root.resizable(0,0)
def randomname(event):
    wb = load_workbook("hello.xlsx")
    ws = wb.active
    rangeline = ws["A2": "A19"]
    name = []
    for items in rangeline:
        for subitems in items:
            name.append(subitems.value)
    computer_action = random.choice(name)
    print("the computer randomly chose: " + computer_action)
    updateDisplay(computer_action)

def updateDisplay(abc):
    displayVariable.set(abc)

button_1 = Button(root, text = "Random Name")
button_1.bind("<Button-1>", randomname)
button_1.pack()
displayVariable = StringVar()
displayLabel = Label(root, textvariable = displayVariable)
displayLabel.pack()

mainloop()